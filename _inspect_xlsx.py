from openpyxl import load_workbook
wb = load_workbook(r'c:\GitHub\mkdocsfirst\photo\纹样数据.xlsx')
ws = wb.active
print('rows', ws.max_row, 'cols', ws.max_column)
for r in range(1, 4):
    print(r, [ws.cell(r, c).value for c in range(1, 5)])
